"""Fonctions ou classes utiles à plusieurs packages."""

import abc
from pathlib import Path
import pandas as pd
import geopandas as gpd
from pandas import DataFrame
from pandas.api.types import is_categorical
from openpyxl.utils.cell import get_column_letter
from .conf import chemins


class Donnée(abc.ABC):
    """Classe abstraite pour conteneur de données sauvegardées dans un fichier.

    Attributs:
        filename: nom du fichier de sauvegarde des données, avec l'extension.
        path: chemin de sauvegarde.
        data: table de données.
    """

    filename: str = ""
    save_path_parent: Path = chemins["transformed"]

    def __init__(self, geo: bool = False) -> None:
        """Charge la table ou la construit et la sauvegarde.

        Args:
            geo: si vrai, charge le fichier parquet avec geopandas.
        """
        self.save_path = self.save_path_parent / self.filename
        suffix = self.save_path.suffix
        try:
            if suffix == ".parquet":
                self.data = (
                    gpd.read_parquet(self.save_path)
                    if geo
                    else pd.read_parquet(self.save_path)
                )
            elif suffix == ".csv":
                self.data = pd.read_csv(self.save_path)
            else:
                raise NotImplementedError(
                    f"{suffix} est un type de fichier non implémenté dans Donnée."
                )
        except FileNotFoundError:
            print(
                f"{self.save_path} n'existe pas. Construction et sauvegarde de la table."
            )
            self.data = self.cstr_base()
            if suffix == ".parquet":
                self.data.to_parquet(self.save_path)
            if suffix == ".csv":
                self.data.to_csv(self.save_path)

    @abc.abstractmethod
    def cstr_base(self) -> DataFrame:
        """Construit la base."""

    @abc.abstractproperty
    def fichier_brut(self) -> DataFrame:
        """Fichier d'origine."""


def xlsx_write_adjust(
    writer: pd.ExcelWriter, table: DataFrame, sheet_name: str, debug: bool = False
) -> None:
    """Écrit le DataFrame sur le fichier Excel avec mise en forme.

    Ajuste la largeur des colonnes et fige les volets.
    BestFit pour la largeur des colonnes ne fonctionne pas pour tous les types de données.

    Arguments:
        writer: doit utiliser l'engine openpyxl
        table: table à sauver
        sheet_name: nom de la feuille du tableur

    Exemple:
        with pd.ExcelWriter("test.xslx") as xl:
            xlsx_write_adjust(xl, df, sheet_name="feuille1")
    """
    pd.options.io.excel.xlsx.writer = "openpyxl"
    table.to_excel(writer, sheet_name=sheet_name)
    worksheet = writer.sheets[sheet_name]
    is_multi = table.columns.nlevels > 1
    try:
        table = table.reset_index()
    except TypeError:  # on essaie d'enlever les catégories
        if is_multi:
            for lev, col in enumerate(table.columns.levels):
                if is_categorical(col):
                    table.columns.set_levels(col.astype(str), level=lev, inplace=True)
        else:
            table.columns = table.columns.astype(str)
        table = table.reset_index()
    for idx, col in enumerate(table):  # loop through all columns
        series = table[col]
        len_name = (
            max(len(str(c)) for c in series.name) if is_multi else len(str(series.name))
        )
        length_max = series.astype(str).map(len).max()
        if debug:
            print(series.name, length_max, len_name)
        max_len = max((length_max, len_name)) + 1  # adding a little extra space
        worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len
        # worksheet.set_column(idx, idx, max_len)  # version xlsxwriter
    worksheet.freeze_panes = get_column_letter(table.columns.nlevels + 1) + str(
        table.index.nlevels + 1
    )
    # version xlsxwriter
    # worksheet.freeze_panes(df.columns.nlevels, df.index.nlevels)
